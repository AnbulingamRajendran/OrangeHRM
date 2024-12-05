import openpyxl
from openpyxl.styles import PatternFill


def row_count(file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    return sheet.max_row


def col_count(file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file_name, sheet_name, r, c):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    return sheet.cell(r, c).value


def write_data(file_name, sheet_name, r, c, data):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    sheet.cell(r, c).value = data
    workbook.save(file_name)


def fill_green(file_name, sheet_name, r, c):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    greenfill = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sheet.cell(r, c).fill = greenfill
    workbook.save(file_name)


def fill_red(file_name, sheet_name, r, c):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    redfill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    sheet.cell(r, c).fill = redfill
    workbook.save(file_name)
